import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

files = [
    r'C:\Users\Lenovo\Downloads\Web back office_Login&Dashboard v1.0.0.xlsx',
    r'C:\Users\Lenovo\Downloads\Web back office_Content management v1.0.0.xlsx',
]

for path in files:
    print(f"\n{'='*60}")
    print(f"FILE: {path}")
    print('='*60)
    wb = openpyxl.load_workbook(path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n--- SHEET: {sheet_name} ---")
        for row in ws.iter_rows(max_row=200, values_only=True):
            if any(cell is not None for cell in row):
                print('\t'.join(str(c) if c is not None else '' for c in row))
